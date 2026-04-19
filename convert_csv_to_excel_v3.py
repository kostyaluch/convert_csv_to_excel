import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re
import html
import threading
import queue
import json
import time

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_FILE_SIZE_MB = 200
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
EXCEL_MAX_ROWS = 1_048_576  # Maximum rows per Excel sheet (including header)

DEFAULT_HEADER_MAP = {
    "goods_article": "Артикул",
    "goods_category_title": "Назва категорії",
    "goods_docket": "Краткое описание",
    "goods_docket_uk": "Короткий опис (ua)",
    "goods_status": "Статус",
    "goods_state": "Стан",
    "goods_title": "Название",
    "goods_title_uk": "Назва (ua)",
}

HEADER_MAP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "header_map.json")

LIGHT_GREEN = "#edf7eb"
DARK_GREEN = "#3d9942"
BUTTON_GREEN = "#6dbf67"
CONVERT_GREEN = "#2e7d32"
CARD_BG = "#ffffff"
HEADER_BG = "#2e7d32"
HEADER_FG = "#ffffff"
BORDER_COLOR = "#b8e0b5"
FONT = ("Segoe UI", 11)
SMALL_FONT = ("Segoe UI", 9)
TITLE_FONT = ("Segoe UI", 15, "bold")

# ---------------------------------------------------------------------------
# Header map persistence
# ---------------------------------------------------------------------------

def load_header_map():
    if os.path.exists(HEADER_MAP_FILE):
        try:
            with open(HEADER_MAP_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return dict(DEFAULT_HEADER_MAP)


def save_header_map(header_map):
    with open(HEADER_MAP_FILE, "w", encoding="utf-8") as f:
        json.dump(header_map, f, ensure_ascii=False, indent=2)

# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def clean_cell_value(value):
    if pd.isnull(value):
        return value
    if isinstance(value, (int, float)):
        return value
    txt = str(value)
    for _ in range(5):
        new_txt = html.unescape(txt)
        if new_txt == txt:
            break
        txt = new_txt
    # BUG-07: replace escaped apostrophe first, then strip control characters only
    txt = txt.replace("\\'", "'")
    for ch in ["\r\n", "\n", "\r", "\t", "\f", "\v"]:  # BUG-05: removed "\s" and "\\"
        txt = txt.replace(ch, "")
    txt = re.sub(r',(?=[^\s])', '; ', txt)
    txt = re.sub(' +', ' ', txt)
    txt = txt.strip()
    return txt


def unique_excel_path(path):
    """Return *path* unchanged if it does not exist, otherwise append _1, _2 … until unique."""
    if not os.path.exists(path):
        return path
    name, ext = os.path.splitext(path)
    counter = 1
    while True:
        candidate = f"{name}_{counter}{ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def format_worksheet(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[1].height = 50
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(15, min(50, max_length * 1.2))


def save_dataframe_to_excel(df, excel_path):
    """Save *df* to *excel_path*.

    If the DataFrame exceeds EXCEL_MAX_ROWS - 1 data rows the output is
    automatically split into *_part1.xlsx*, *_part2.xlsx*, ... files.
    Returns a list of paths that were actually written.
    BUG-09: output path is made unique so existing files are never overwritten silently.
    BUG-08: workbook is always closed via finally to prevent descriptor leaks.
    """
    saved_files = []
    usable_rows = EXCEL_MAX_ROWS - 1  # one row reserved for the header

    if len(df) <= usable_rows:
        out_path = unique_excel_path(excel_path)
        df.to_excel(out_path, index=False)
        wb = load_workbook(out_path)
        try:
            format_worksheet(wb.active)
            wb.save(out_path)
        finally:
            wb.close()
        saved_files.append(out_path)
    else:
        base, ext = os.path.splitext(excel_path)
        part = 1
        start = 0
        while start < len(df):
            chunk = df.iloc[start:start + usable_rows]
            part_path = unique_excel_path(f"{base}_part{part}{ext}")
            chunk.to_excel(part_path, index=False)
            wb = load_workbook(part_path)
            try:
                format_worksheet(wb.active)
                wb.save(part_path)
            finally:
                wb.close()
            saved_files.append(part_path)
            start += usable_rows
            part += 1

    return saved_files


def _ua_variant_of(col):
    """Return the expected ``(ua)`` variant name for *col*.

    Handles two naming patterns:
    * ``'Name'``          → ``'Name (ua)'``
    * ``'Name;ID'``       → ``'Name (ua);ID'``   (attribute columns with a numeric ID)
    """
    m = re.match(r'^(.*?)(;\d+)$', col)
    if m:
        return m.group(1) + " (ua)" + m.group(2)
    return col + " (ua)"


def pair_ua_columns(df):
    """Reorder *df* columns so that every ``'X (ua)'`` column immediately follows
    the matching base column ``'X'``, but **only** when both columns contain
    digits in their name (i.e. they carry a numeric attribute ID such as
    ``';20775'``).  Columns without digits (plain service headers like
    ``'Назва (ua)'``) are left in their original positions.

    Columns that have no ``(ua)`` counterpart keep their original position.
    ``(ua)`` columns that have no matching base are appended at the end in their
    original relative order.  Returns (reordered_df, n_pairs) where *n_pairs* is
    the number of base↔(ua) pairs that were actually moved.
    """
    cols = list(df.columns)
    col_set = set(cols)

    # Build set of (ua) columns that are eligible for pairing:
    # they must contain digits (carry an ID) and their base column must also exist.
    ua_paired = set()
    for col in cols:
        if re.search(r'\d', col):
            ua_var = _ua_variant_of(col)
            if ua_var in col_set and ua_var != col:
                # col is a base; mark its ua variant as paired
                ua_paired.add(ua_var)

    ordered = []
    placed = set()
    n_pairs = 0

    for col in cols:
        if col in placed:
            continue
        if col in ua_paired:
            continue  # skip eligible (ua) columns; placed right after their base
        ordered.append(col)
        placed.add(col)
        # Only attempt pairing when the base column itself has digits
        if re.search(r'\d', col):
            ua_var = _ua_variant_of(col)
            if ua_var in ua_paired and ua_var not in placed:
                ordered.append(ua_var)
                placed.add(ua_var)
                n_pairs += 1

    # append any (ua) columns whose base was not found
    for col in cols:
        if col not in placed:
            ordered.append(col)

    return df[ordered], n_pairs

# ---------------------------------------------------------------------------
# Tooltip helper
# ---------------------------------------------------------------------------

class ToolTip:
    """Simple hover tooltip for any widget."""

    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self._tip = None
        widget.bind("<Enter>", self._show)
        widget.bind("<Leave>", self._hide)

    def _show(self, _event=None):
        if self._tip:
            return
        x = self.widget.winfo_rootx() + 24
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self._tip = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tk.Label(
            tw, text=self.text, background="#fffde7", relief="solid",
            borderwidth=1, font=SMALL_FONT, padx=6, pady=3
        ).pack()

    def _hide(self, _event=None):
        if self._tip:
            self._tip.destroy()
            self._tip = None

# ---------------------------------------------------------------------------
# Header-map editor dialog
# ---------------------------------------------------------------------------

class HeaderMapDialog(tk.Toplevel):
    """Modal dialog for viewing and editing the header-rename dictionary."""

    def __init__(self, parent, header_map):
        super().__init__(parent)
        self.title("Словник заголовків")
        self.configure(bg=LIGHT_GREEN)
        self.resizable(True, True)
        self.geometry("520x420")
        self.result = None
        self._rows = []
        self._canvas = None

        ttk.Label(self, text="Оригінальна назва → Нова назва", font=FONT,
                  background=LIGHT_GREEN).pack(pady=(10, 4))

        # Scrollable area for rows
        container = ttk.Frame(self)
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        canvas = tk.Canvas(container, bg=LIGHT_GREEN, highlightthickness=0)
        self._canvas = canvas
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        self.scroll_frame = ttk.Frame(canvas)
        self.scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        for orig, renamed in header_map.items():
            self._add_row(orig, renamed)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=6)
        ttk.Button(btn_frame, text="Додати слово", command=lambda: self._add_row()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Видалити останній", command=self._delete_last).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Зберегти", command=self._save).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frame, text="Скасувати", command=self.destroy).pack(side=tk.LEFT, padx=4)

        self.grab_set()
        self.wait_window()

    def _add_row(self, orig="", renamed=""):
        row_frame = ttk.Frame(self.scroll_frame)
        row_frame.pack(fill=tk.X, pady=2)
        orig_var = tk.StringVar(value=orig)
        renamed_var = tk.StringVar(value=renamed)
        orig_entry = ttk.Entry(row_frame, textvariable=orig_var, width=24)
        orig_entry.pack(side=tk.LEFT, padx=(4, 2))
        ttk.Label(row_frame, text="→").pack(side=tk.LEFT, padx=2)
        ttk.Entry(row_frame, textvariable=renamed_var, width=24).pack(side=tk.LEFT, padx=(2, 4))
        self._rows.append((row_frame, orig_var, renamed_var))
        if self._canvas and not orig and not renamed:
            self.update_idletasks()
            self._canvas.yview_moveto(1.0)
            orig_entry.focus_set()

    def _delete_last(self):
        if self._rows:
            row_frame, _, _ = self._rows.pop()
            row_frame.destroy()

    def _save(self):
        result = {}
        for _, orig_var, renamed_var in self._rows:
            orig = orig_var.get().strip()
            renamed = renamed_var.get().strip()
            if orig:
                result[orig] = renamed
        self.result = result
        save_header_map(result)
        self.destroy()

# ---------------------------------------------------------------------------
# Main application
# ---------------------------------------------------------------------------

class CsvToExcelConverterApp:
    def __init__(self, master):
        self.master = master
        master.title("Конвертер CSV у Excel")
        master.geometry("740x680")
        master.configure(bg=HEADER_BG)
        master.resizable(True, True)

        # Center on screen
        master.update_idletasks()
        sw, sh = master.winfo_screenwidth(), master.winfo_screenheight()
        master.geometry(f"740x680+{(sw - 740) // 2}+{(sh - 680) // 2}")

        # ── Styles ────────────────────────────────────────────────────────
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TButton', font=FONT, background=BUTTON_GREEN, foreground='black',
                        borderwidth=0, focusthickness=2, focuscolor=DARK_GREEN, padding=(10, 6))
        style.map('TButton',
                  background=[('active', DARK_GREEN), ('disabled', '#cccccc')],
                  foreground=[('disabled', '#888888')])
        style.configure('Convert.TButton', font=("Segoe UI", 13, "bold"),
                        background=CONVERT_GREEN, foreground='white',
                        borderwidth=0, padding=(24, 11))
        style.map('Convert.TButton',
                  background=[('active', '#1b5e20'), ('disabled', '#aaaaaa')],
                  foreground=[('disabled', '#dddddd')])
        style.configure('TLabel', background=LIGHT_GREEN, font=FONT)
        style.configure('TFrame', background=LIGHT_GREEN)
        style.configure('TLabelframe', background=LIGHT_GREEN,
                        bordercolor=BORDER_COLOR, relief='groove')
        style.configure('TLabelframe.Label', background=LIGHT_GREEN,
                        font=("Segoe UI", 10, "bold"), foreground=DARK_GREEN)
        style.configure('TCheckbutton', background=LIGHT_GREEN, font=FONT)
        style.configure(
            'Card.TCheckbutton',
            background=LIGHT_GREEN,
            foreground=DARK_GREEN,
            font=("Segoe UI", 11, "bold"),
            padding=(2, 6)
        )
        style.map(
            'Card.TCheckbutton',
            foreground=[('active', '#1b5e20'), ('selected', '#1b5e20')],
            background=[('active', LIGHT_GREEN), ('selected', LIGHT_GREEN)]
        )
        style.configure('TRadiobutton', background=LIGHT_GREEN, font=FONT)
        style.configure('Vertical.TScrollbar', background=LIGHT_GREEN,
                        arrowcolor=DARK_GREEN, troughcolor=BORDER_COLOR)
        style.configure('TProgressbar', background=DARK_GREEN,
                        troughcolor=BORDER_COLOR, borderwidth=0, thickness=18)

        # ── Header bar ────────────────────────────────────────────────────
        header_bar = tk.Frame(master, bg=HEADER_BG, pady=11)
        header_bar.pack(fill=tk.X, side=tk.TOP)

        tk.Label(
            header_bar, text="📊  Конвертер CSV у Excel",
            font=TITLE_FONT, bg=HEADER_BG, fg=HEADER_FG
        ).pack(side=tk.LEFT, padx=20)

        gear_btn = tk.Button(
            header_bar, text="⚙", font=("Segoe UI", 14),
            bg=HEADER_BG, fg=HEADER_FG,
            activebackground="#1b5e20", activeforeground=HEADER_FG,
            relief="flat", cursor="hand2", borderwidth=0,
            command=self.open_header_map_dialog
        )
        gear_btn.pack(side=tk.RIGHT, padx=14)
        ToolTip(gear_btn, "Словник заголовків")

        # ── Main content ──────────────────────────────────────────────────
        main_frame = ttk.Frame(master, padding=(18, 14, 18, 10))
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ── Files card ────────────────────────────────────────────────────
        files_lf = ttk.LabelFrame(main_frame, text="  📁  Файли  ", padding=(10, 6))
        files_lf.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(
            files_lf,
            text="Оберіть один або декілька CSV-файлів для конвертації",
            font=SMALL_FONT, foreground="#777777"
        ).pack(anchor="w", pady=(0, 5))

        list_frame = ttk.Frame(files_lf)
        list_frame.pack(fill=tk.X)

        scrollbar_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        self.files_listbox = tk.Listbox(
            list_frame, height=5, font=("Segoe UI", 10),
            bg=CARD_BG, borderwidth=1, relief="solid",
            highlightcolor=DARK_GREEN, selectbackground=BUTTON_GREEN,
            activestyle='none', yscrollcommand=scrollbar_y.set
        )
        scrollbar_y.config(command=self.files_listbox.yview)
        self.files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        file_btn_bar = ttk.Frame(files_lf)
        file_btn_bar.pack(fill=tk.X, pady=(7, 0))

        self.select_button = ttk.Button(
            file_btn_bar, text="📂  Вибрати файли", command=self.select_files
        )
        self.select_button.pack(side=tk.LEFT, padx=(0, 8))

        self.clear_button = ttk.Button(
            file_btn_bar, text="🗑  Очистити", command=self.clear_files
        )
        self.clear_button.pack(side=tk.LEFT)

        self.files_count_label = ttk.Label(
            file_btn_bar, text="Файлів не вибрано",
            font=SMALL_FONT, foreground="#888888"
        )
        self.files_count_label.pack(side=tk.RIGHT, padx=4)

        # ── Options card ──────────────────────────────────────────────────
        opts_lf = ttk.LabelFrame(main_frame, text="  ⚙  Налаштування  ", padding=(10, 8))
        opts_lf.pack(fill=tk.X, pady=(0, 10))

        self.clean_var = tk.StringVar(value="clean")
        mode_frame = ttk.Frame(opts_lf)
        mode_frame.pack(anchor="w", pady=(0, 5))
        ttk.Label(mode_frame, text="Форматування:").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(
            mode_frame, text="Очищений", variable=self.clean_var, value="clean"
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Radiobutton(
            mode_frame, text="Оригінальний", variable=self.clean_var, value="original"
        ).pack(side=tk.LEFT)

        self.delete_csv_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            opts_lf, text="Видалити CSV-файли після конвертації",
            variable=self.delete_csv_var, style='Card.TCheckbutton', cursor='hand2'
        ).pack(anchor="w")

        self.pair_ua_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opts_lf, text="Групувати (ua)-стовпці поруч із базовими",
            variable=self.pair_ua_var, style='Card.TCheckbutton', cursor='hand2'
        ).pack(anchor="w")

        # ── Convert button ────────────────────────────────────────────────
        self.convert_button = ttk.Button(
            main_frame, text="🚀  Конвертувати у Excel",
            command=self.convert_files, style='Convert.TButton'
        )
        self.convert_button.pack(pady=(2, 10))

        # ── Progress ──────────────────────────────────────────────────────
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 8))

        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_bar = ttk.Progressbar(
            progress_frame, variable=self.progress_var,
            maximum=100, mode='determinate'
        )
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))

        self.progress_label = ttk.Label(
            progress_frame, text="0%", font=SMALL_FONT,
            foreground=DARK_GREEN, width=5, anchor="e"
        )
        self.progress_label.pack(side=tk.RIGHT)

        # ── Log section ───────────────────────────────────────────────────
        ttk.Label(
            main_frame, text="Журнал процесу",
            font=("Segoe UI", 11, "bold"), foreground=DARK_GREEN
        ).pack(anchor="w", pady=(0, 3))

        self.log_text = scrolledtext.ScrolledText(
            main_frame, height=8, font=("Segoe UI", 10),
            bg=CARD_BG, borderwidth=1, relief="solid", wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.configure(state='disabled')

        # ── State ─────────────────────────────────────────────────────────
        self.selected_files = []
        self.header_map = load_header_map()
        self._queue = queue.Queue()

    # ------------------------------------------------------------------
    # File selection
    # ------------------------------------------------------------------

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Виберіть CSV-файли для конвертації",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]  # TECH-06
        )
        if files:
            self.selected_files = list(files)
            self.files_listbox.delete(0, tk.END)
            for f in self.selected_files:
                self.files_listbox.insert(tk.END, f)
            n = len(self.selected_files)
            self.files_count_label.configure(
                text=f"{n} файл{'ів' if n != 1 else ''} вибрано", foreground=DARK_GREEN
            )
            self.log("Файли вибрано. Готові до конвертації.")

    def clear_files(self):
        self.selected_files = []
        self.files_listbox.delete(0, tk.END)
        self.files_count_label.configure(text="Файлів не вибрано", foreground="#888888")

    # ------------------------------------------------------------------
    # Header map dialog
    # ------------------------------------------------------------------

    def open_header_map_dialog(self):
        dlg = HeaderMapDialog(self.master, self.header_map)
        if dlg.result is not None:
            self.header_map = dlg.result
            self.log("Словник заголовків оновлено та збережено.")

    # ------------------------------------------------------------------
    # Conversion (TECH-01: runs in background thread)
    # ------------------------------------------------------------------

    def convert_files(self):
        if not self.selected_files:
            messagebox.showwarning("Увага", "Будь ласка, виберіть файли для конвертації.")
            return

        # TECH-02: warn about large files before starting
        large_files = [
            (f, os.path.getsize(f) / (1024 * 1024))
            for f in self.selected_files
            if os.path.getsize(f) > MAX_FILE_SIZE_BYTES
        ]
        if large_files:
            file_list = "\n".join(
                f"• {os.path.basename(f)} ({s:.0f} МБ)" for f, s in large_files
            )
            if not messagebox.askyesno(
                "Великі файли",
                f"Наступні файли перевищують {MAX_FILE_SIZE_MB} МБ:\n{file_list}\n\n"
                "Конвертація може тривати довго. Продовжити?"
            ):
                return

        self.select_button.configure(state='disabled')
        self.convert_button.configure(state='disabled')
        self.progress_var.set(0.0)
        self._queue = queue.Queue()

        thread = threading.Thread(target=self._conversion_worker, daemon=True)
        thread.start()
        self.master.after(100, self._poll_queue)

    def _conversion_worker(self):
        files = self.selected_files
        total = len(files)
        success_count = 0
        error_count = 0
        delete_csv = self.delete_csv_var.get()
        header_map = self.header_map
        clean_mode = self.clean_var.get() == "clean"
        pair_ua = self.pair_ua_var.get()

        for idx, file_path in enumerate(files):
            try:
                self._queue.put({"type": "log", "text": f"📂 Читаю файл: {file_path}"})
                t0 = time.time()

                df = pd.read_csv(file_path, encoding='utf-8', sep=',')
                self._queue.put({"type": "log",
                                 "text": f"   Рядків: {len(df)}, стовпців: {len(df.columns)}"})

                if clean_mode:
                    self._queue.put({"type": "log", "text": "   Очищення даних..."})
                    df = df.applymap(clean_cell_value)

                # Rename headers according to the dictionary
                rename_map = {k: v for k, v in header_map.items() if k in df.columns}
                if rename_map:
                    df = df.rename(columns=rename_map)

                # Pair (ua) columns immediately after their base counterpart
                if pair_ua:
                    df, n_pairs = pair_ua_columns(df)
                    if n_pairs:
                        self._queue.put({"type": "log",
                                         "text": f"   Згруповано пар (ua): {n_pairs}"})

                excel_base_path = os.path.splitext(file_path)[0] + '.xlsx'
                self._queue.put({"type": "log", "text": "   Збереження у Excel..."})

                saved_files = save_dataframe_to_excel(df, excel_base_path)
                for saved in saved_files:
                    self._queue.put({"type": "log",
                                     "text": f"✅ Збережено: {os.path.basename(saved)}"})

                elapsed = time.time() - t0
                self._queue.put({"type": "log", "text": f"   Час обробки: {elapsed:.1f} с"})

                if delete_csv:
                    os.remove(file_path)
                    self._queue.put({"type": "log",
                                     "text": f"🗑️ CSV видалено: {os.path.basename(file_path)}"})

                success_count += 1
            except Exception as e:
                self._queue.put({"type": "log",
                                 "text": f"❌ Помилка у файлі {os.path.basename(file_path)}: {e}"})
                error_count += 1

            self._queue.put({"type": "progress", "value": (idx + 1) / total * 100})

        self._queue.put({"type": "done", "success": success_count, "error": error_count})

    def _poll_queue(self):
        try:
            for _ in range(20):  # process at most 20 messages per cycle to keep GUI responsive
                msg = self._queue.get_nowait()
                msg_type = msg.get("type")
                if msg_type == "log":
                    self.log(msg["text"])
                elif msg_type == "progress":
                    val = msg["value"]
                    self.progress_var.set(val)
                    self.progress_bar["value"] = val
                    self.progress_label.configure(text=f"{val:.0f}%")
                elif msg_type == "done":
                    self._on_conversion_done(msg["success"], msg["error"])
                    return
        except queue.Empty:
            pass
        self.master.after(100, self._poll_queue)

    def _on_conversion_done(self, success_count, error_count):
        self.progress_var.set(100.0)
        self.progress_bar["value"] = 100.0
        self.progress_label.configure(text="100%")
        self.log(f"\nЗавершено! Успішно: {success_count}, з помилкою: {error_count}")
        self.select_button.configure(state='normal')
        self.convert_button.configure(state='normal')
        messagebox.showinfo("Готово", f"Успішно конвертовано: {success_count}\nЗ помилкою: {error_count}")

    # ------------------------------------------------------------------
    # Logging
    # ------------------------------------------------------------------

    def log(self, message):
        self.log_text.configure(state='normal')
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state='disabled')


if __name__ == "__main__":
    root = tk.Tk()
    app = CsvToExcelConverterApp(root)
    root.mainloop()
